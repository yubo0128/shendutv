import xlsxwriter
import openpyxl
import copy
import logging

# 配置日志记录
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# 定义常量
BASIC_DATA = ["一、", "二、", "三、", "四、", "五、", "六、", "七、", "八、", "九、", "十、"]
# 固有匹配 如果匹配到就取消前面的换行符号
MATCHED_VALUE = ["如发现以上症状，应及时去正规医院就医，以免耽误病情。",
                 "一旦确诊，应积极配合医生治疗，遵医嘱，按疗程，科学治，以免病情加重。",
                 "远离这些因素，养成良好的生活习惯，健康的生活方式是健康最有力的保障。",
                 "特别提醒，要到正规医院就医，切勿相信偏方，小广告等，以免错过治疗时机。",
                 "治疗期间要相信医生，不恐惧，不焦虑，不迷信，科学治疗和良好心态是治疗成功的有力保障。"]

'''
    返回比较的值
    @:param str1: str, 比较值一
    @:param str2: str, 比较值二
    @:return  随机生成的不重复序列返回1.0 是全匹配  0 是全不匹配
'''


def jaccard_similarity(str1, str2):
    set1 = set(str1)
    set2 = set(str2)
    intersection = len(set1.intersection(set2))
    union = len(set1.union(set2))
    return intersection / union


'''
    返回截取的数据
    @:param c: str, 截取的数据
    @:param string: BASIC_DATA, 常量
    @:return  随机生成的不重复序列返回1.0 是全匹配  0 是全不匹配
'''


def computes(c, string):
    rag = 0
    out = c[0:2]
    if out in string:
        rag = 2
    if len(c) > 15:
        couA = c[rag:15]
    else:
        couA = c[rag:10]
    return couA


'''
    @:param sh 读取的第一个表格
    @:param arr: list, 声明的空列表
    @:return  无
'''


# 处理数据
def main(sh, arr):
    # 读取excel并组装数据
    for cases in list(sh.rows)[1:]:
        a = cases[4].value
        b = cases[5].value
        # print(b)
        # 把数据转换成A（左列）和B（右列）
        arr.append({
            "a": "\n".join([s for s in a.split("\n") if s]),
            "b": "\n".join([s for s in b.split("\n") if s])
        })
    # print(arr)
    index = 0
    for itemss in arr:
        s = itemss['a']
        n = itemss["b"]
        # if index == 0:
        #     index += 1
        #     continue
        # 循环常量
        for st in BASIC_DATA:
            logging.info('查询当前循环的数据' + st)
            if st in s:
                i = s.index(st)
                sub_str = s[int(i):i + 15]
                # 第一种情况 带一 二 三的情况
                if sub_str in n:  # 带BASIC_DATAg的情况
                    wz = n.index(sub_str)
                    ss = list(n)
                    ss.insert(wz, '\n')  # 在索引位置插入一个空格
                    n = ''.join(ss)
                else:  # 不带一二三的情况
                    sub_str = s[int(i):i + 15]
                    print(sub_str)
                    sp = sub_str.split("。")
                    sub_str2 = s[int(i) + 2: (i + 2 + len(sp[0]))]
                    # 第二种情况左列存在一二三的情况
                    if sub_str in n:
                        wz = n.index(sub_str)
                        character = n[0:wz]
                        character_list = [s for s in character.split("\n") if s]
                        if len(character_list) != 1:
                            wz = wz - 2
                        ss = list(n)
                        ss.insert(wz, '\n' + st)  # 在索引位置插入一个空格
                        n = ''.join(ss)
                        continue
                    # 第三种情况不处理
                    elif sub_str2 in n:
                        logging.info("是什么" + n)
                        logging.error('第三种情况')
                        # wz = n.index(sub_str2)
                        # ss = list(n)
                        # print(ss)
                        # ss.insert(wz, '\n')  # 在索引位置插入一个空格
                        # n = ''.join(ss)
                        # print(n)
            else:
                logging.error('不存在')
        nl = [s for s in n.split("\n") if s]
        print(nl)
        itemss["b"] = "\n".join(nl)
        index += 1

    '''
    第四种情况处理第一行和第二行分行情况
    '''
    for i in arr:
        # 把字符串拆分成数组
        a = i['a'].split("\n")
        b = i['b'].split("\n")
        bs = i['b']
        if len(a) != len(b):
            if a[1] in b[0]:
                wz = bs.index(a[1])
                ss = list(bs)
                ss.insert(wz, '\n')  # 在索引5的位置插入一个空格
                bs = ''.join(ss)
                print(bs)
                i['b'] = bs
            else:
                if a[0] in b[0]:
                    wz = bs.index(a[0])
                    ss = list(bs)
                    ss.insert(wz + len(a[0]), '\n')
                    bs = ''.join(ss)
                    i['b'] = bs
            a = [s for s in i['b'].split("\n") if s]
            i['b'] = '\n'.join(a)
        else:
            if a[0] in b[0]:
                ss = list(bs)
                ss.insert(int(len(a[0])), '\n')  # 在索引5的位置插入一个空格
                bs = ''.join(ss)
                i['b'] = bs
            else:
                print("这种情况不处理")
            a = [s for s in i['b'].split("\n") if s]
            i['b'] = '\n'.join(a)

    '''
    第五种情况有些带了一二三有些没带说明前面的字符串不配需拼接上一二三
    '''
    for item in arr:
        a = item["a"].split("\n")
        b = item["b"].split("\n")
        for aa in a:
            for i in range(0, len(b)):
                # 截取后面10个字符或者15个字符超过95% 就加编号
                deep_copy_list = copy.deepcopy(aa)
                cou_a = computes(deep_copy_list, BASIC_DATA)
                cou_b = computes(b[i], BASIC_DATA)
                if jaccard_similarity(cou_a, cou_b) > 0.70:
                    sindex = aa[0:2]
                    if sindex in BASIC_DATA:
                        if sindex in b[i]:
                            b[i] = b[i]
                        else:
                            b[i] = sindex + b[i]
        '''
        新增功能如果遇到任意一句就取消前面换行
        '''
        for i in range(0, len(b)):
            for v in MATCHED_VALUE:
                if v in b[i]:
                    if i > 0:
                        b[i-1] = b[i-1] + b[i]
                        b[i] = ""
            # if b[i] in MATCHED_VALUE:
            #     if i > 0:
            #         b[i-1] = b[i-1] + b[i]
            #         b[i] = ""
        item["b"] = "\n".join([s for s in b if s])

'''
# 导出和飘红
@:param arr: list, s数据
@:param output_name 导出的局对路径
'''


def output(arr, output_name):
    # 创建工作簿对象
    workbook = openpyxl.Workbook()
    # 选取默认的活动表格（sheet）
    sheet = workbook.active
    workbook.save(output_name)
    workbook = xlsxwriter.Workbook(output_name)
    worksheet = workbook.add_worksheet()

    # 定义格式
    red_format = workbook.add_format({'color': 'red'})
    normal_format = workbook.add_format()  # 默认的格式
    worksheet.write(0, 0, 'XQ版本正文')
    worksheet.write(0, 1, 'XQ版本扩增正文-20240307')

    # 循环
    for item in range(0, len(arr)):
        z = arr[item]
        a = z['a'].split("\n")
        b = z['b'].split("\n")
        final_data = []
        id = 0
        index_id = []
        for aStr in a:
            for i in range(0, len(b)):
                if jaccard_similarity(aStr, b[i]) > 0.85:
                    # 必须要得到B列的key值
                    index_id.append(i)
                    break
            id += 1
        # 组装数据并设置飘红
        for k in range(0, len(b)):
            if k in index_id:
                final_data.append(normal_format)
                if (k + 1) == len(b):
                    final_data.append(b[k])
                else:
                    final_data.append(b[k] + "\n")
            else:
                final_data.append(red_format)
                if (k + 1) == len(b):
                    final_data.append(b[k])
                else:
                    final_data.append(b[k] + "\n")
        print(final_data)
        # 输出并飘红
        # # 第一个参数是行号，第二个参数是列号，后面的参数是交替的文本和格式
        worksheet.write_rich_string(int(item + 1), 0, "\n".join(a), normal_format, " ")
        worksheet.write_rich_string(int(item + 1), 1, " ", *final_data)
    # 关闭工作簿，释放内存
    workbook.close()


if __name__ == '__main__':
    # 全数据测试
    # wb = openpyxl.load_workbook("/Users/yubo/Desktop/未命名文件夹/多囊卵巢综合征-XQ版扩增拼接+正文-361条-20240313.xlsx")
    # wb = openpyxl.load_workbook("/Users/yubo/Desktop/未命名文件夹/多囊卵巢综合征-XQ版扩增拼接+正文-361条-20240318.xlsx")
    # wb = openpyxl.load_workbook("/Users/yubo/Desktop/未命名文件夹/20240319/阴道炎XQ扩展版拼接+正文-195条-20240319.xlsx")
    wb = openpyxl.load_workbook("/Users/yubo/Desktop/未命名文件夹/20240319/盆腔炎XQ扩增版本拼接+正文-1328条-20240319.xlsx")
    # 部分数据测试
    # wb = openpyxl.load_workbook("./工作簿115.xlsx")
    sh = wb.worksheets[0]
    arr = []
    # 处理数据
    main(sh, arr)
    # 导出数据 参数一数据 参数二是导出文件名称
    output(arr, "盆腔炎XQ扩增版本拼接+正文-1328条-20240319.xlsx")
