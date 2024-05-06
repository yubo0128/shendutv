import xlsxwriter
import openpyxl
import copy
import re
import sys
import datetime

# 获取当前日期
now = datetime.datetime.now()

# 定义常量
BASIC_DATA = ["一、", "二、", "三、", "四、", "五、", "六、", "七、", "八、", "九、", "十、"]
# 固有匹配 如果匹配到就取消前面的换行符号
MATCHED_VALUE = ["如发现以上症状，应及时去正规医院就医，以免耽误病情。",
                 "一旦确诊，应积极配合医生治疗，遵医嘱，按疗程，科学治，以免病情加重。",
                 "远离这些因素，养成良好的生活习惯，健康的生活方式是健康最有力的保障。",
                 "特别提醒，要到正规医院就医，切勿相信偏方，小广告等，以免错过治疗时机。",
                 "治疗期间要相信医生，不恐惧，不焦虑，不迷信，科学治疗和良好心态是治疗成功的有力保障。",
                 "诊断疾病请去正规医院，进行科学的检查，早确诊，早治疗。"]

'''
    返回比较的值
    @:param str1: str, 比较值一
    @:param str2: str, 比较值二
    @:return  随机生成的不重复序列返回1.0 是全匹配  0 是全不匹配
'''


def jaccard_similarity(str1, str2):
    set1 = set(str1)
    set2 = set(str2)
    intersection = len(set1.intersection(set2))
    union = len(set1.union(set2))
    return intersection / union


'''
    返回截取的数据
    @:param c: str, 截取的数据
    @:param string: BASIC_DATA, 常量
    @:return  随机生成的不重复序列返回1.0 是全匹配  0 是全不匹配
'''


def computes(c, string):
    rag = 0
    out = c[0:2]
    if out in string:
        rag = 2
    if len(c) > 30:
        couA = c[rag:len(c)]
    else:
        couA = c[rag:10]
    return couA


def computes2(c, string):
    rag = 0
    out = c[0:2]
    if out in string:
        rag = 2
    if len(c) > 15:
        couA = c[rag:15]
    else:
        couA = c[rag:10]
    return couA


'''
根据value获取key值
'''


def getKey(value):
    # 根据BASIC_DATA获取key
    for i in range(len(BASIC_DATA)):
        if BASIC_DATA[i] == value:
            return i


'''
    @:param sh 读取的第一个表格
    @:param arr: list, 声明的空列表
    @:return  无
'''


# 处理数据
def main(sh, arr):
    # 读取excel并组装数据
    for cases in list(sh.rows)[1:]:
        a = cases[4].value
        b = cases[5].value
        b = b.replace("^l", "").replace("^l", "").replace("^l", "").replace(" ", "").replace(" ", "")
        a = a.replace("^l", "").replace("^l", "").replace("^l", "").replace(" ", "").replace(" ", "")
        # print(b)
        # 把数据转换成A（左列）和B（右列）
        arr.append({
            "a": "\n".join([s for s in a.split("\n") if s]),
            "b": "\n".join([s for s in b.split("\n") if s])
        })

    # print(arr)
    index = 0
    for itemss in arr:
        s = itemss['a']
        n = itemss["b"]
        # if index == 0:
        #     index += 1
        #     continue
        # 循环常量
        for st in BASIC_DATA:
            if st in s:
                i = s.index(st)
                sub_str = s[int(i):i + 15]
                # 第一种情况 带一 二 三的情况
                if sub_str in n:  # 带BASIC_DATAg的情况
                    wz = n.index(sub_str)
                    ss = list(n)
                    ss.insert(wz, '\n')  # 在索引位置插入一个空格
                    n = ''.join(ss)
                else:  # 不带一二三的情况
                    sub_str = s[int(i):i + 15]
                    # print(sub_str)
                    sp = sub_str.split("。")
                    sub_str2 = s[int(i) + 2: (i + 2 + len(sp[0]))]
                    # 第二种情况左列存在一二三的情况
                    if sub_str in n:
                        wz = n.index(sub_str)
                        character = n[0:wz]
                        character_list = [s for s in character.split("\n") if s]
                        if len(character_list) != 1:
                            wz = wz - 2
                        ss = list(n)
                        ss.insert(wz, '\n' + st)  # 在索引位置插入一个空格
                        n = ''.join(ss)
                        continue
                    # 第三种情况不处理
                    # elif sub_str2 in n:
                    #     print("")
                    # wz = n.index(sub_str2)
                    # ss = list(n)
                    # print(ss)
                    # ss.insert(wz, '\n')  # 在索引位置插入一个空格
                    # n = ''.join(ss)
                    # print(n)
            # else:
                # print("")
        nl = [s for s in n.split("\n") if s]
        # print(nl)
        itemss["b"] = "\n".join(nl)
        index += 1
    '''
    第四种情况处理第一行和第二行分行情况
    '''
    for i in arr:
        # 把字符串拆分成数组
        a = i['a'].split("\n")
        b = i['b'].split("\n")
        bs = i['b']
        if len(a) != len(b):
            # print(a)
            # print("输出22222：")
            if len(a) > 24:
                if a[1] in b[0]:
                    wz = bs.index(a[1])
                    ss = list(bs)
                    ss.insert(wz, '\n')  # 在索引5的位置插入一个空格
                    bs = ''.join(ss)
                    # print(bs)
                    i['b'] = bs
                else:
                    if a[0] in b[0]:
                        wz = bs.index(a[0])
                        ss = list(bs)
                        ss.insert(wz + len(a[0]), '\n')
                        bs = ''.join(ss)
                        i['b'] = bs
                nasc = [s for s in i['b'].split("\n") if s]
                i['b'] = '\n'.join(nasc)
        else:
            if a[0] in b[0]:
                ss = list(bs)
                ss.insert(int(len(a[0])), '\n')  # 在索引5的位置插入一个空格
                bs = ''.join(ss)
                i['b'] = bs
            # else:
            #     print("")
            ansdos = [s for s in i['b'].split("\n") if s]
            i['b'] = '\n'.join(ansdos)
    # print("aaaaaaaaaaaaaaaaa开始")
    # print(itemss["a"])
    # print("aaaaaaaaaaaaaaaaa结束")
    # print("bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb")
    # print(itemss["b"])
    # print("bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb")

    # 2024年04月10日新增第二行换行规则
    for item in arr:
        a = item['a'].split("\n")
        b = item["b"].split("\n")
        # print("输出内容")
        # print(b)
        # # 左边第二列要去匹配右边第二列 如果存在就换行
        if len(a) > 1:
            if a[1] in b[1]:
                ss = list(b[1])
                ss.insert(int(len(a[1])), '\n')  # 在索引5的位置插入一个空格
                # print("输出内容112"+ str(ss))
                b[1] = ''.join(ss)
                # print("存在匹配222")
        item["b"] = '\n'.join(b)
    # print("删除删除删除删除")
    # print(arr)
    # print("删除删除删除删除")
    '''
    第五种情况有些带了一二三有些没带说明前面的字符串不配需拼接上一二三
    '''
    for item in arr:
        a = item["a"].split("\n")
        b = item["b"].split("\n")
        index = 0
        amax = None
        amaxIndex = 0
        amaxIndexNumber = 0
        bmax = None
        bmaxIndex = 0
        bmaxIndexNumber = 0
        # print("22222222222222222222222222222")
        # print(b)
        # print("22222222222222222222222222222")
        for aa in a:
            for pp in BASIC_DATA:
                if pp in aa:
                    # print("vvvvc" + pp)
                    amax = pp
                    amaxIndex = amaxIndex + 1
                    # 得到行号A的行号
                    # print("得到第几行" + str(index))
                    amaxIndexNumber = index
            for i in range(0, len(b)):
                # 截取后面10个字符或者15个字符超过95% 就加编号
                if len(aa) > 10:
                    deep_copy_list = copy.deepcopy(aa)
                else:
                    # print("执行合并小于十个字符的")
                    deep_copy_list = copy.deepcopy(aa + a[index + 1])
                cou_a = computes(deep_copy_list, BASIC_DATA)
                cou_b = computes(b[i], BASIC_DATA)
                # print("打印出值AAAA" + cou_a)
                # print("打印出值BBBB" + cou_b)
                # print("打印出匹配的值" + str(jaccard_similarity(cou_a, cou_b)))
                # 这里调小了会造成数据前面匹配上了一、 二、 三、的情况

                cou_a_2 = computes2(deep_copy_list, BASIC_DATA)
                cou_b_2 = computes2(b[i], BASIC_DATA)
                p = False
                # 这里需要验证一下开头两位是否存在一二三编码的情况 如果不存在才执行
                sindex = aa[0:2]
                bindex = b[i][0:2]
                rbh = ",".join(BASIC_DATA)
                if bindex not in rbh:
                    if (jaccard_similarity(cou_a, cou_b) > 0.75) and i > 0:
                        p = True
                        if sindex in BASIC_DATA:
                            if sindex in b[i]:
                                b[i] = b[i]
                            else:
                                # print("输出一二三的情况" + b[i][0:2])
                                if b[i][0:2] != "1、":
                                    b[i] = sindex + b[i]
                    if p == False and jaccard_similarity(cou_a_2, cou_b_2) > 0.75 and i > 0:
                        if sindex in BASIC_DATA:
                            if sindex in b[i]:
                                b[i] = b[i]
                            else:
                                # print("输出一二三的情况")
                                if b[i][0:2] != "1、":
                                    b[i] = sindex + b[i]

            index = index + 1

        '''
        新增功能如果遇到任意一句就取消前面换行
        1、修复匹配值不正确问题
        '''
        for i in range(0, len(b)):
            for pp in BASIC_DATA:
                if pp in b[i]:
                    # print("111111" + str(bmaxIndex))
                    bmax = pp
                    bmaxIndex = bmaxIndex + 1
                    bmaxIndexNumber = i
            for v in MATCHED_VALUE:
                if v == b[i]:
                    if i > 0:
                        b[i - 1] = b[i - 1] + b[i]
                        b[i] = ""
            # if b[i] in MATCHED_VALUE:
            #     if i > 0:
            #         b[i-1] = b[i-1] + b[i]
            #         b[i] = ""
        # print("A最大值" + str(amax))
        # print("B最大值" + str(bmax))
        # print("A最大值11" + str(amaxIndex))
        # print("B最大值22" + str(bmaxIndex))
        # print("A得到a的行号" + str(amaxIndexNumber))
        # print("B得到b的行号" + str(bmaxIndexNumber))
        # if len(a) == len(b):
        #     keyIndex = getKey(amax)
        #     print("得到Key" + str(keyIndex))
        #     for key in range(amaxIndexNumber, 0, -1):
        #         if keyIndex < 0:
        #             break
        #         print("测试测试" + b[key][0:2])
        #         print("得到Key" + b[key])
        #         index_index = BASIC_DATA[keyIndex]
        #         if index_index not in b[key][0:2] and BASIC_DATA[keyIndex + 1] not in b[key][0:2] and BASIC_DATA[
        #             keyIndex - 1] not in b[key][0:2]:
        #             print("得到Key111" + b[key])
        #             b[key] = BASIC_DATA[keyIndex] + b[key]
        #             print("1111")
        #             print(BASIC_DATA[keyIndex])
        #             print("1111")
        #         keyIndex = keyIndex - 1
        # 先从最大值往前遍历
        keyIndex = getKey(bmax)
        for key in range(bmaxIndexNumber + 1, 0, -1):
            # print("测试测试" + str(key - 1))
            if keyIndex == None or keyIndex < 0:
                break
            # print("得到字符" + str(keyIndex))
            # print("得到字符" + BASIC_DATA[keyIndex])
            # print(b[key - 1])
            index_index = BASIC_DATA[keyIndex]
            # print("AAAAAAAAA")
            jieQu = b[key - 1][0:2]
            # print(BASIC_DATA[keyIndex + 1])
            # print(BASIC_DATA[keyIndex - 1])
            # print("AAAAAAAAA")
            if index_index not in jieQu and BASIC_DATA[keyIndex + 1] not in jieQu and BASIC_DATA[
                keyIndex - 1] not in jieQu:
                # print("分到了232322323"+ BASIC_DATA[keyIndex])
                # print(b[key - 1])
                if b[key - 1] != "" and b[key - 1][0:2] != "1、" and b[key - 1][0:2] != "2、" and b[key - 1][0:2] != "3、":
                    b[key - 1] = BASIC_DATA[keyIndex] + b[key - 1]
            keyIndex = keyIndex - 1

        # 处理左边的序号大于右边的序号就要在右边加
        if amax is not None and bmax is not None and amax <= bmax:
            # 获取到得到A的最大值
            Aindex = getKey(amax)
            Bindex = getKey(bmax)
            # print("A最大值" + str(Aindex))
            # print("B最大值" + str(Bindex))
            compute = Aindex - Bindex
            # print("相差多少为" + str(compute))
            # print(b)
            if compute > 0:
                ljj = 1
                for i in range(Bindex + 1, Aindex + 1):
                    # print("是什么值"+str(bmaxIndexNumber + ljj))
                    # print(b[bmaxIndexNumber + ljj])
                    # print("大于0" + BASIC_DATA[i])
                    b[bmaxIndexNumber + +ljj] = BASIC_DATA[i] + b[bmaxIndexNumber + +ljj]
                    ljj = ljj + 1
        item["b"] = "\n".join([s for s in b if s])
        # print("bbbb")
        # print(item["b"])


'''
# 导出和飘红
@:param arr: list, s数据
@:param output_name 导出的局对路径
'''


def output(arr, output_name, otherArr):
    # 创建工作簿对象
    workbook = openpyxl.Workbook()
    # 选取默认的活动表格（sheet）
    sheet = workbook.active
    workbook.save(output_name)
    workbook = xlsxwriter.Workbook(output_name)
    worksheet = workbook.add_worksheet()

    # 定义格式
    red_format = workbook.add_format({'color': 'red'})
    normal_format = workbook.add_format()  # 默认的格式

    worksheet.write(0, 0, '序号')
    worksheet.write(0, 1, 'title')
    worksheet.write(0, 2, '封面编号')
    worksheet.write(0, 3, 'XQ版本A')
    worksheet.write(0, 4, 'XQ版本正文')
    worksheet.write(0, 5, 'XQ版本扩增正文-' + str(now.year) + "-" + str(now.month) + "-" + str(now.day))
    worksheet.write(0, 6, '模块1')
    worksheet.write(0, 7, '模块2')
    worksheet.write(0, 8, '模块3')
    worksheet.write(0, 9, '模块4')
    worksheet.write(0, 10, '模块5')
    worksheet.write(0, 11, '模块6')
    worksheet.write(0, 12, '模块7')
    worksheet.write(0, 13, '模块8')
    worksheet.write(0, 14, '模块9')
    worksheet.write(0, 15, '模块10')
    worksheet.write(0, 16, '模块11')
    worksheet.write(0, 17, '模块12')
    worksheet.write(0, 18, '模块13')
    worksheet.write(0, 19, '模块14')
    worksheet.write(0, 20, '模块15')
    # worksheet.write(0, 0, 'XQ版本正文')
    # worksheet.write(0, 1, 'XQ版本扩增正文-' + str(now.year) + "-" + str(now.month) + "-" + str(now.day))

    # 循环
    for item in range(0, len(arr)):
        z = arr[item]
        a = z['a'].split("\n")
        b = z['b'].split("\n")
        final_data = []
        id = 0
        index_id = []
        for aStr in a:
            for i in range(0, len(b)):
                if jaccard_similarity(aStr, b[i]) > 0.85:
                    # 必须要得到B列的key值
                    index_id.append(i)
                    # break
            id += 1
        # 组装数据并设置飘红
        for k in range(0, len(b)):
            if k in index_id:
                final_data.append(normal_format)
                if (k + 1) == len(b):
                    final_data.append(b[k])
                else:
                    final_data.append(b[k] + "\n")
            else:
                final_data.append(red_format)
                if (k + 1) == len(b):
                    final_data.append(b[k])
                else:
                    final_data.append(b[k] + "\n")
        # print(final_data)
        # worksheet.write_rich_string(int(item + 1), 0, "\n".join(a), normal_format, " ")
        # worksheet.write_rich_string(int(item + 1), 1, " ", *final_data)
        # 输出并飘红
        # 第一个参数是行号，第二个参数是列号，后面的参数是交替的文本和格式
        for i in range(0, len(otherArr[item])):
            worksheet.write_rich_string(int(item + 1), i, otherArr[item][i], normal_format, " ")
        if otherArr[item][4] == "4":
            worksheet.write_rich_string(int(item + 1), 4, "\n".join(a), normal_format, " ")
        if otherArr[item][5] == "5":
            worksheet.write_rich_string(int(item + 1), 5, " ", *final_data)
    # 关闭工作簿，释放内存
    workbook.close()


if __name__ == '__main__':
    input_file = sys.argv[1]
    output_file = sys.argv[2]
    # 全数据测试
    wb = openpyxl.load_workbook(input_file)
    sh = wb.worksheets[0]
    arr = []
    # 处理数据
    main(sh, arr)

    otherArr = []
    # 记录
    for cases in list(sh.rows)[1:]:
        childArr = []
        for i in range(6, len(cases)):
            if cases[i].value is not None:
                childArr.append(cases[i].value)
        otherArr.append([cases[0].value, cases[1].value, cases[2].value, cases[3].value, "4", "5"] + childArr)
    output(arr, output_file, otherArr)
    print(output_file)
