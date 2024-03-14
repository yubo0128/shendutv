import xlsxwriter
import openpyxl
import copy


# 返回比较的值
# 1.0 是全匹配  0 是全不匹配
def jaccard_similarity(str1, str2):
    set1 = set(str1)
    set2 = set(str2)
    intersection = len(set1.intersection(set2))
    union = len(set1.union(set2))
    return intersection / union


# 计算数据
def computes(c, string):
    rag = 0
    out = c[0:2]
    if out in string:
        rag = 2
    print(rag)
    if len(c) > 20:
        couA = c[rag:20]
    else:
        couA = c[rag:10]
    return couA


wb = openpyxl.load_workbook("工作簿115.xlsx")
sh = wb.worksheets[0]
arr = []
for cases in list(sh.rows)[1:]:
    a = cases[4].value
    b = cases[5].value
    # print(b)
    arr.append({
        "a": "\n".join([s for s in a.split("\n") if s]),
        "b": "\n".join([s for s in b.split("\n") if s])
    })
# print(arr)
index = 0
string = ["一、", "二、", "三、", "四、", "五、", "六、", "七、", "八、", "九、", "十、"]
for itemss in arr:
    s = itemss['a']
    n = itemss["b"]
    # if index == 0:
    #     index += 1
    #     continue
    for st in string:
        print("查询当前循环的数据：" + st)
        if st in s:
            i = s.index(st)
            sub_str = s[int(i):i + 15]
            # 第一种情况
            if sub_str in n:  # 带string的情况
                print("查询到了111")
                wz = n.index(sub_str)
                ss = list(n)
                ss.insert(wz, '\n')  # 在索引位置插入一个空格
                n = ''.join(ss)
                print("-------ssss---")
                print(n)
                print("-------ssss---")
            else:  # 不带一二三的情况
                sub_str = s[int(i) + 2:i + 15]
                print("查询出2222222")
                print(sub_str)
                sp = sub_str.split("。")
                sub_str2 = s[int(i) + 2: (i + 2 + len(sp[0]))]
                print("输出sub_str2sub_str2sub_str2")
                print(sub_str2)
                print("输出sub_str2sub_str2sub_str2")
                print(n)
                print(sub_str in n)
                print("第几次" + str(index))
                # 第二种情况
                if sub_str in n:
                    wz = n.index(sub_str)
                    print(wz)
                    print("数字")
                    ss = list(n)
                    ss.insert(wz, '\n' + st)  # 在索引位置插入一个空格
                    n = ''.join(ss)
                    print("查询到了")
                    print("-------查询到了---")
                    print(n)
                    print("-------查询到了---")
                    continue
                # 第三种情况
                elif sub_str2 in n:
                    print("是什么" + n)
                    print("走截取")
                    # wz = n.index(sub_str2)
                    # ss = list(n)
                    # print(ss)
                    # ss.insert(wz, '\n')  # 在索引位置插入一个空格
                    # n = ''.join(ss)
                    # print(n)
        else:
            print("不存在")
    nl = [s for s in n.split("\n") if s]
    print("000000000000000000")
    print(nl)
    itemss["b"] = "\n".join(nl)
    index += 1
# exit()
print(arr)
print("111111111111111111111111111111111111111111")
# 第四种情况
for i in arr:
    a = i['a'].split("\n")
    b = i['b'].split("\n")
    bs = i['b']
    print("进入进入进入进入进入进入")
    if len(a) != len(b):
        print("------------_________--------")
        print("BBBBB" + b[0])
        print("AAAAA" + a[1])
        if a[1] in b[0]:
            wz = bs.index(a[1])
            ss = list(bs)
            ss.insert(wz, '\n')  # 在索引5的位置插入一个空格
            bs = ''.join(ss)
            print("1111111")
            print(bs)
            print("------------______2222___--------")
            i['b'] = bs
        else:
            print("进入进入进入进入进入进入2222222222222")
            if a[0] in b[0]:
                print("走包含了111111")
                print(len(a[0]))
                wz = bs.index(a[0])
                ss = list(bs)
                ss.insert(wz + len(a[0]), '\n')
                bs = ''.join(ss)
                print(bs)
                i['b'] = bs
        print("输出BBBB")
        print(i['b'].split("\n"))
        a = [s for s in i['b'].split("\n") if s]
        i['b'] = '\n'.join(a)
    else:
        if a[0] in b[0]:
            print("AAAAAAAAA1111:" + a[1])
            print("BBBBBBBBB1111:" + b[0])
            print("CCCCCCCCC:" + str(len(a[0])))
            ss = list(bs)
            ss.insert(int(len(a[0])), '\n')  # 在索引5的位置插入一个空格
            bs = ''.join(ss)
            print("1111111")
            print(bs)
            print("------------______2222___--------")
            i['b'] = bs
            print("8888888888888888888888888")
        else:
            print("99999999999999999999999999999999999")
        a = [s for s in i['b'].split("\n") if s]
        i['b'] = '\n'.join(a)
print("222222222222222222222222222222222222222222222222")
# 第五种情况有些带了一二三有些没带说明前面的字符串不配需拼接上一二三
for item in arr:
    a = item["a"].split("\n")
    b = item["b"].split("\n")
    for aa in a:
        for i in range(0, len(b)):
            # 截取后面10个字符或者15个字符超过95% 就加编号
            deep_copy_list = copy.deepcopy(aa)
            cou_a = computes(deep_copy_list, string)
            cou_b = computes(b[i], string)
            print("查询条数")
            print("AAA值：" + cou_a)
            print("BBB值：" + cou_b)
            print(jaccard_similarity(cou_a, cou_b))
            if jaccard_similarity(cou_a, cou_b) > 0.70:
                sindex = aa[0:2]
                print(sindex)
                if sindex in string:
                    b[i] = sindex + b[i]
                    print("进；来了")
    item["b"] = "\n".join(b)


# print(arr)
# exit()

def output(arr):
    # 创建工作簿对象
    workbook = openpyxl.Workbook()
    # 选取默认的活动表格（sheet）
    sheet = workbook.active
    filename = "output111.xlsx"
    workbook.save(filename)
    workbook = xlsxwriter.Workbook(filename)
    worksheet = workbook.add_worksheet()

    # 定义格式
    red_format = workbook.add_format({'color': 'red'})
    normal_format = workbook.add_format()  # 默认的格式
    worksheet.write(0, 0, 'XQ版本正文')
    worksheet.write(0, 1, 'XQ版本扩增正文-20240307')
    #

    for item in range(0, len(arr)):
        z = arr[item]
        a = z['a'].split("\n")
        b = z['b'].split("\n")
        print(b)
        variables = {}
        p = []
        if len(a) == len(b) or len(a) > len(b):
            for index in range(0, len(a)):
                variable_name = f'a{index}'
                # print(a[index])
                print(index)
                bb = ""
                aa = ""
                try:
                    aa = a[index].replace(" ", "").replace(" ", "").replace(" ", "").replace("\n", "").replace("\n\r",
                                                                                                               "")
                    if index < len(b):
                        bb = b[index].replace(" ", "").replace(" ", "").replace(" ", "").replace("\n", "").replace(
                            "\n\r",
                            "")
                    if aa == bb:
                        print(aa)
                        print("等于111111")
                        print(bb)
                        variables[variable_name] = False
                        p.append(normal_format)
                        p.append(bb + "\n")
                    else:
                        print(aa)
                        print("不等于11111111")
                        print(bb)
                        variables[variable_name] = True
                        p.append(red_format)
                        p.append(bb + "\n")
                except IndexError as e:
                    print("抛出异常：" + e)
                    # p.append(normal_format)
                    # p.append(bb + "\n")
        elif len(a) < len(b):
            # a循环匹配b 如果b
            print("a循环匹配b 如果b")
            id = 0
            index_id = []
            for aStr in a:
                for i in range(0, len(b)):
                    print(jaccard_similarity(aStr, b[i]))
                    if jaccard_similarity(aStr, b[i]) > 0.85:
                        print("进入记录成功匹配的")
                        print("aaa:"+aStr)
                        print("bbb:"+b[i])
                        index_id.append(id)
                        break
                id += 1
            print(index_id)
            for k in range(0, len(b)):
                if k in index_id:
                    p.append(normal_format)
                    p.append(b[k] + "\n")
                else:
                    p.append(red_format)
                    p.append(b[k] + "\n")
            print(p)

            # for index in range(0, len(b)):
            #     variable_name = f'a{index}'
            #     # print(a[index])
            #     print(index)
            #     # print(b[index])
            #     try:
            #         aa = ""
            #         if index < len(a):
            #             aa = a[index].replace(" ", "").replace(" ", "").replace(" ", "").replace("\n", "").replace(
            #                 "\n\r",
            #                 "")
            #         bb = b[index].replace(" ", "").replace(" ", "").replace(" ", "").replace("\n", "").replace("\n\r",
            #                                                                                                    "")
            #         print("ppppppppppppppppp")
            #         print(aa)
            #         print(bb)
            #         print("ppppppppppppppppp")
            #
            #         if aa == bb:
            #             print(aa)
            #             print("等于2222222222")
            #             print(bb)
            #             variables[variable_name] = False
            #             p.append(normal_format)
            #             p.append(bb + "\n")
            #         else:
            #             print(aa)
            #             print("不等于22222222")
            #             print(bb)
            #             variables[variable_name] = True
            #             p.append(red_format)
            #             p.append(bb + "\n")
            #     except IndexError as e:
            #         # p.append(normal_format)
            #         # p.append(bb + "\n")
            #         print("错误信息：" + e)
        # # 第一个参数是行号，第二个参数是列号，后面的参数是交替的文本和格式
        worksheet.write_rich_string(int(item + 1), 0, "\n".join(a), normal_format, " ")
        worksheet.write_rich_string(int(item + 1), 1, " ", *p)
    # 关闭工作簿，释放内存
    workbook.close()


output(arr)
# workbook = xlsxwriter.Workbook('./工作簿5.xlsx')
# worksheet = workbook.add_worksheet()

